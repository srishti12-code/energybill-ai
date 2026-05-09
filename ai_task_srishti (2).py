import streamlit as st
import openpyxl
from PIL import Image

st.title("Electricity Bill Automation")

uploaded_file = st.file_uploader(
    "Upload Bill Image",
    type=["jpg", "jpeg", "png"]
)

if uploaded_file:

    image = Image.open(uploaded_file)

    st.image(image, caption="Uploaded Bill", use_container_width=True)

    image.save("bill.jpg")
    extracted_text="Bill text Extraction disable"

    st.subheader("Extracted Text")
    st.write(extracted_text)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Extracted Text"
    ws["A2"] = extracted_text

    wb.save("Final_Output.xlsx")

    with open("Final_Output.xlsx", "rb") as file:
        st.download_button(
            label="Download Excel File",
            data=file,
            file_name="Final_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
